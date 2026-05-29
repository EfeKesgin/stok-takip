import csv
import io
from flask import Flask, render_template, request, redirect, url_for, flash, Response
from models import db, Product, Transaction
from sqlalchemy import func
from datetime import datetime, timedelta

app = Flask(__name__)
app.config['SECRET_KEY'] = 'stok_takip_erp_secret_2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///stok.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()

@app.route('/')
def dashboard():
    total_products = Product.query.count()
    total_items = db.session.query(func.sum(Product.quantity)).scalar() or 0
    
    all_prods = Product.query.all()
    total_value = sum(p.quantity * p.current_price for p in all_prods)
    
    low_stock_products = Product.query.filter(Product.quantity < 10).all()
    recent_transactions = Transaction.query.order_by(Transaction.date.desc()).limit(5).all()
    
    category_data_query = db.session.query(Product.category, func.sum(Product.quantity)).group_by(Product.category).all()
    categories = [row[0] for row in category_data_query]
    category_counts = [row[1] or 0 for row in category_data_query]
    
    # Grafik verileri (Son 7 gün)
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_txs_all = Transaction.query.filter(Transaction.date >= seven_days_ago).order_by(Transaction.date.asc()).all()
    
    time_range = request.args.get('range', '24h')
    now = datetime.now()
    
    labels = []
    chart_in = []
    chart_out = []
    
    if time_range == '24h':
        labels = [(now - timedelta(hours=i)).strftime('%H:00') for i in range(23, -1, -1)]
        chart_in = [0]*24
        chart_out = [0]*24
        start_time = now - timedelta(hours=24)
        for t in Transaction.query.filter(Transaction.date >= start_time).all():
            l_str = t.date.strftime('%H:00')
            if l_str in labels:
                idx = labels.index(l_str)
                if t.type == 'IN': chart_in[idx] += t.quantity
                else: chart_out[idx] += t.quantity
    elif time_range == '1week':
        labels = [(now - timedelta(days=i)).strftime('%d.%m') for i in range(6, -1, -1)]
        chart_in = [0]*7
        chart_out = [0]*7
        start_time = now - timedelta(days=7)
        for t in Transaction.query.filter(Transaction.date >= start_time).all():
            l_str = t.date.strftime('%d.%m')
            if l_str in labels:
                idx = labels.index(l_str)
                if t.type == 'IN': chart_in[idx] += t.quantity
                else: chart_out[idx] += t.quantity
    elif time_range == '1month':
        labels = [(now - timedelta(days=i)).strftime('%d.%m') for i in range(29, -1, -1)]
        chart_in = [0]*30
        chart_out = [0]*30
        start_time = now - timedelta(days=30)
        for t in Transaction.query.filter(Transaction.date >= start_time).all():
            l_str = t.date.strftime('%d.%m')
            if l_str in labels:
                idx = labels.index(l_str)
                if t.type == 'IN': chart_in[idx] += t.quantity
                else: chart_out[idx] += t.quantity
    elif time_range == '3months':
        labels = [(now - timedelta(days=i)).strftime('%d.%m') for i in range(89, -1, -1)]
        chart_in = [0]*90
        chart_out = [0]*90
        start_time = now - timedelta(days=90)
        for t in Transaction.query.filter(Transaction.date >= start_time).all():
            l_str = t.date.strftime('%d.%m')
            if l_str in labels:
                idx = labels.index(l_str)
                if t.type == 'IN': chart_in[idx] += t.quantity
                else: chart_out[idx] += t.quantity
    elif time_range == '1year':
        labels = []
        chart_in = [0]*12
        chart_out = [0]*12
        for i in range(11, -1, -1):
            m_date = now - timedelta(days=i*30)
            labels.append(m_date.strftime('%Y-%m'))
        start_time = now - timedelta(days=365)
        for t in Transaction.query.filter(Transaction.date >= start_time).all():
            l_str = t.date.strftime('%Y-%m')
            if l_str in labels:
                idx = labels.index(l_str)
                if t.type == 'IN': chart_in[idx] += t.quantity
                else: chart_out[idx] += t.quantity
    else:
        start_time = now - timedelta(days=7) # fallback safe
                
    total_stock_cost = 0.0
    total_revenue = 0.0
    total_profit = 0.0
    total_discount_loss = 0.0
    
    period_txs = Transaction.query.filter(Transaction.date >= start_time).all()
    for t in period_txs:
        p = t.product
        if p:
            if t.type == 'IN':
                total_stock_cost += t.quantity * getattr(t, 'unit_purchase_price', p.purchase_price)
            elif t.type == 'OUT':
                t_price = getattr(t, 'unit_price', p.current_price)
                t_purch = getattr(t, 'unit_purchase_price', p.purchase_price)
                t_loss = getattr(t, 'unit_discount_loss', 0.0)
                
                total_revenue += t.quantity * t_price
                total_profit += t.quantity * (t_price - t_purch)
                total_discount_loss += t.quantity * t_loss

    low_stock_products = Product.query.filter(Product.quantity <= 5).order_by(Product.quantity.asc()).limit(8).all()
    
    best_sellers = db.session.query(
        Product.name,
        func.sum(Transaction.quantity).label('total_sold')
    ).join(Transaction, Product.id == Transaction.product_id).filter(
        Transaction.type == 'OUT',
        Transaction.date >= start_time
    ).group_by(Product.id).order_by(func.sum(Transaction.quantity).desc()).limit(5).all()
    
    return render_template('index.html', 
                           total_products=total_products,
                           total_items=total_items,
                           total_value=total_value,
                           recent_transactions=recent_transactions,
                           categories=categories,
                           category_counts=category_counts,
                           chart_dates=labels,
                           chart_in=chart_in,
                           chart_out=chart_out,
                           current_range=time_range,
                           total_stock_cost=total_stock_cost,
                           total_revenue=total_revenue,
                           total_profit=total_profit,
                           total_discount_loss=total_discount_loss,
                           low_stock_products=low_stock_products,
                           best_sellers=best_sellers)

@app.route('/products')
def products():
    products = Product.query.order_by(Product.id.asc()).all()
    return render_template('products.html', products=products)

@app.route('/logs')
def logs():
    # En son işlemleri (son 1000 kayıt) azalan tarih sırasına göre getir
    transactions = Transaction.query.order_by(Transaction.date.desc()).limit(1000).all()
    return render_template('logs.html', transactions=transactions)

@app.route('/export_logs')
def export_logs():
    transactions = Transaction.query.order_by(Transaction.date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Islem_Gecmisi"
    
    headers = ["Tarih Saat", "İşlem Yönü", "SKU", "Ürün Adı", "Kategori", "Adet", "İşlem Anı Etiket Fiyatı", "İşlem Anı Maliyet", "Net Kâr/Zarar", "İndirim Kaybı"]
    ws.append(headers)
    
    for t in transactions:
        net_profit = ""
        discount_loss = ""
        cost = ""
        price = ""
        if t.type == 'OUT':
            cost = t.unit_purchase_price if t.unit_purchase_price else 0
            price = t.unit_price if t.unit_price else 0
            net_profit = (price - cost) * t.quantity
            discount_loss = (t.unit_discount_loss * t.quantity) if t.unit_discount_loss else 0
            
        row = [
            t.date.strftime("%Y-%m-%d %H:%M:%S"),
            "GİRİŞ" if t.type == 'IN' else "ÇIKIŞ (SATIŞ)",
            t.product.sku,
            t.product.name,
            t.product.category,
            t.quantity,
            price,
            cost,
            net_profit,
            discount_loss
        ]
        ws.append(row)
        
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return Response(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=islem_gecmisi.xlsx'}
    )

@app.route('/api/products', methods=['GET'])
def api_get_products():
    products = Product.query.all()
    return jsonify([{'sku': p.sku, 'name': p.name, 'quantity': p.quantity, 'price': p.current_price} for p in products])

@app.route('/scanner')
def scanner():
    return render_template('scanner.html')

@app.route('/api/product/<sku>')
def api_product(sku):
    p = Product.query.filter_by(sku=sku).first()
    if p:
        return {'success': True, 'id': p.id, 'name': p.name, 'price': p.current_price, 'quantity': p.quantity}
    return {'success': False, 'message': 'Ürün bulunamadı'}

@app.route('/campaigns')
def campaigns():
    products = Product.query.all()
    
    active_campaigns = []
    available_products = []
    
    for p in products:
        if p.current_price < p.price:
            active_campaigns.append(p)
        else:
            available_products.append(p)
            
        # Son kullanma tarihi geçmiş indirimleri veritabanından kalıcı temizle
        if p.discount_end_date and p.discount_end_date <= datetime.now():
            p.discount_type = None
            p.discount_value = 0.0
            p.campaign_buy_x = 0
            p.campaign_pay_y = 0
            p.discount_end_date = None
            db.session.commit()
            
    return render_template('campaigns.html', active_campaigns=active_campaigns, available_products=available_products)

@app.route('/add_product', methods=['POST'])
def add_product():
    sku = request.form.get('sku')
    name = request.form.get('name')
    category = request.form.get('category')
    quantity = int(request.form.get('quantity', 0))
    price = float(request.form.get('price', 0.0))
    purchase_price = float(request.form.get('purchase_price', 0.0))
    
    new_product = Product(sku=sku, name=name, category=category, quantity=quantity, price=price, purchase_price=purchase_price)
    db.session.add(new_product)
    db.session.commit()
    
    if quantity > 0:
        transaction = Transaction(product_id=new_product.id, type='IN', quantity=quantity)
        db.session.add(transaction)
        db.session.commit()
        
    flash('Ürün başarıyla eklendi.', 'success')
    return redirect(url_for('products'))

@app.route('/set_discount/<int:id>', methods=['POST'])
def set_discount(id):
    product = Product.query.get_or_404(id)
    d_type = request.form.get('discount_type')
    d_val = request.form.get('discount_value', type=float, default=0.0)
    days = request.form.get('discount_days', type=int, default=7)
    buy_x = request.form.get('buy_x', type=int, default=0)
    pay_y = request.form.get('pay_y', type=int, default=0)
    
    if d_type in ['PERCENT', 'AMOUNT'] and d_val > 0 and days > 0:
        product.discount_type = d_type
        product.discount_value = d_val
        product.campaign_buy_x = 0
        product.campaign_pay_y = 0
        product.discount_end_date = datetime.now() + timedelta(days=days)
        flash('Kampanya tanımlandı.', 'success')
    elif d_type == 'BUY_X_PAY_Y' and buy_x > 0 and pay_y > 0 and days > 0:
        product.discount_type = d_type
        product.discount_value = 0.0
        product.campaign_buy_x = buy_x
        product.campaign_pay_y = pay_y
        product.discount_end_date = datetime.now() + timedelta(days=days)
        flash(f'{buy_x} Al {pay_y} Öde kampanyası başlatıldı.', 'success')
    else:
        product.discount_type = None
        product.discount_value = 0.0
        product.campaign_buy_x = 0
        product.campaign_pay_y = 0
        product.discount_end_date = None
        flash('Eksik/Hatalı giriş, kampanya sıfırlandı.', 'warning')
        
    db.session.commit()
    return redirect(url_for('campaigns'))

@app.route('/edit_product/<int:id>', methods=['POST'])
def edit_product(id):
    product = Product.query.get_or_404(id)
    product.sku = request.form.get('sku')
    product.name = request.form.get('name')
    product.category = request.form.get('category')
    product.price = float(request.form.get('price', 0.0))
    product.purchase_price = float(request.form.get('purchase_price', 0.0))
    product.quantity = int(request.form.get('quantity', 0))
    
    db.session.commit()
    flash('Ürün bilgileri güncellendi.', 'success')
    return redirect(url_for('products'))

@app.route('/remove_discount/<int:id>')
def remove_discount(id):
    product = Product.query.get_or_404(id)
    product.discount_type = None
    product.discount_value = 0.0
    product.campaign_buy_x = 0
    product.campaign_pay_y = 0
    product.discount_end_date = None
    db.session.commit()
    flash('Kampanya başarıyla iptal edildi.', 'success')
    return redirect(url_for('campaigns'))

@app.route('/delete_product/<int:id>')
def delete_product(id):
    product = Product.query.get_or_404(id)
    Transaction.query.filter_by(product_id=product.id).delete()
    db.session.delete(product)
    db.session.commit()
    flash('Ürün silindi.', 'danger')
    return redirect(url_for('products'))

@app.route('/transaction', methods=['POST'])
def add_transaction():
    product_id = request.form.get('product_id')
    t_type = request.form.get('type')
    qty = int(request.form.get('quantity'))
    
    product = Product.query.get(product_id)
    if not product:
        flash('Ürün bulunamadı!', 'danger')
        return redirect(request.referrer or url_for('products'))
        
    if t_type == 'OUT' and product.quantity < qty:
        flash('Yetersiz stok!', 'danger')
        return redirect(request.referrer or url_for('products'))
        
    u_price = 0.0
    u_purch = product.purchase_price
    u_loss = 0.0

    if t_type == 'IN':
        product.quantity += qty
    elif t_type == 'OUT':
        product.quantity -= qty
        u_price = product.current_price
        u_loss = (product.price - product.current_price) if product.price > product.current_price else 0.0
        
    transaction = Transaction(
        product_id=product.id, 
        type=t_type, 
        quantity=qty,
        unit_price=u_price,
        unit_purchase_price=u_purch,
        unit_discount_loss=u_loss
    )
    db.session.add(transaction)
    db.session.commit()
    
    flash('Stok hareketi eklendi.', 'success')
    return redirect(request.referrer or url_for('products'))

@app.route('/export')
def export_products():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    
    products = Product.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stok Listesi"
    
    headers = ['SKU', 'Ürün Adı', 'Kategori', 'Stok', 'Alış Fiyatı', 'Satış Fiyatı', 'Toplam Satış Değeri']
    ws.append(headers)
    
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    for p in products:
        c_price = p.current_price
        row = [p.sku, p.name, p.category, p.quantity, getattr(p, 'purchase_price', 0.0), c_price, p.quantity * c_price]
        ws.append(row)
        
    for row in ws.iter_rows(min_row=2, max_col=7, max_row=len(products)+1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
            
    widths = [15, 30, 20, 10, 15, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=stok_listesi.xlsx'}
    )

@app.route('/api/import_products', methods=['POST'])
def import_products():
    try:
        data = request.json
        if not data:
            return {'success': False, 'error': 'Veri bulunamadı'}
        
        count = 0
        for row in data:
            sku = str(row.get('SKU') or row.get('Stok Kodu') or row.get('StokKodu') or '').strip()
            name = str(row.get('Name') or row.get('Ürün Adı') or row.get('Ürün') or '').strip()
            
            if not sku or not name:
                continue
                
            qty = int(row.get('Quantity') or row.get('Adet') or row.get('Stok') or 0)
            
            p_val = row.get('Price') or row.get('Satış Fiyatı') or row.get('Fiyat') or 0
            price = float(str(p_val).replace(',', '.') if isinstance(p_val, str) else p_val)
            
            b_val = row.get('Purchase Price') or row.get('Alış Fiyatı') or row.get('Maliyet') or 0
            purchase = float(str(b_val).replace(',', '.') if isinstance(b_val, str) else b_val)
            
            category = str(row.get('Category') or row.get('Kategori') or 'Genel').strip()
            
            p = Product.query.filter_by(sku=sku).first()
            if p:
                p.quantity += qty
                if price > 0: p.price = price
                if purchase > 0: p.purchase_price = purchase
            else:
                p = Product(sku=sku, name=name, category=category, quantity=qty, price=price, purchase_price=purchase)
                db.session.add(p)
                
            count += 1
            
        db.session.commit()
        return {'success': True, 'count': count, 'message': f'{count} adet ürün içe aktarıldı.'}
    except Exception as e:
        return {'success': False, 'error': str(e)}

if __name__ == '__main__':
    app.run(debug=True, port=5000)
